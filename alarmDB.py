import openpyxl as xl
import datetime as dt
import os

alarmFile = 'alarms.xlsx'

if not os.path.exists(alarmFile):
  xl.Workbook().save(alarmFile)

def wb():   
  wb = xl.load_workbook(filename = alarmFile)
  return wb

def getAlarms():
  l1=[]
  for row in wb().active.values:
    l2=[]
    for col in row:
      l2.append(col)
    l1.append(l2)
  return l1

def addAlarm(dtString, name):
  wba = wb()
  dts = dt.datetime.strptime(dtString, "%m/%d/%Y %H:%M:%S")
  row = [wba.active.max_row, dts, name]
  wba.active.append(row)
  wba.save(alarmFile)